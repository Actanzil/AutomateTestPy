import openpyxl
import time
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service

start_time = time.time()
print("Memulai Program")
driver = webdriver.Chrome()
driver.get("https://forms.gle/ueXgvWTQRxAb56St7")

wait = WebDriverWait(driver, 10)
wait.until(EC.visibility_of_element_located((By.CLASS_NAME, 'whsOnd.zHQkBf')))

def fill_form(email, nrp, nama, jenis_kelamin, usia, alamat, golongan, posisi, department, divisi, site, pilihan):

    inputs = driver.find_elements(By.CLASS_NAME, 'whsOnd.zHQkBf')
    time.sleep(1)

    inputs_array = [email]

    for i in range(len(inputs)):
        inputs[i].clear()
        inputs[i].send_keys(inputs_array[i])

    bt_next = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="mG61Hd"]/div[2]/div/div[3]/div[1]/div[1]/div')))
    bt_next.click()

    wait.until(EC.visibility_of_element_located((By.CLASS_NAME, 'whsOnd.zHQkBf')))
    wait.until(EC.visibility_of_element_located((By.CLASS_NAME, 'KHxj8b.tL9Q4c')))
    wait.until(EC.visibility_of_element_located((By.XPATH, '//div[@class="lRwqcd"]//span[text()="Berikutnya"]')))

    inputs = driver.find_elements(By.CLASS_NAME, 'whsOnd.zHQkBf')
    radiobuttons = driver.find_elements(By.CLASS_NAME, 'Od2TWd.hYsg7c')
    textareas = driver.find_element(By.CLASS_NAME, 'KHxj8b.tL9Q4c')
    time.sleep(1)

    inputs_array = [nrp, nama, usia]

    for i in range(len(inputs)):
        inputs[i].clear()
        inputs[i].send_keys(inputs_array[i])

    textareas.clear()
    textareas.send_keys(alamat)

    for radio in radiobuttons:
        if radio.get_attribute("data-value").lower() == jenis_kelamin.lower():
            radio.click()

    bt_next = wait.until(EC.element_to_be_clickable((By.XPATH, '//div[@class="lRwqcd"]//span[text()="Berikutnya"]')))
    bt_next.click()

    wait.until(EC.visibility_of_element_located((By.CLASS_NAME, 'whsOnd.zHQkBf')))
    wait.until(EC.visibility_of_element_located((By.XPATH, '//div[@class="lRwqcd"]//span[text()="Berikutnya"]')))

    inputs = driver.find_elements(By.CLASS_NAME, 'whsOnd.zHQkBf')
    radiobuttons = driver.find_elements(By.CLASS_NAME, 'Od2TWd.hYsg7c')
    time.sleep(1)  # Tunggu sebentar

    inputs_array = [golongan, department, divisi, site]

    for i in range(len(inputs)):
        inputs[i].clear()
        inputs[i].send_keys(inputs_array[i])

    for radio in radiobuttons:
        if radio.get_attribute("data-value").lower() == posisi.lower():
            radio.click()

    bt_next = wait.until(EC.element_to_be_clickable((By.XPATH, '//div[@class="lRwqcd"]//span[text()="Berikutnya"]')))
    bt_next.click()

    wait.until(EC.visibility_of_element_located((By.CLASS_NAME, 'Od2TWd.hYsg7c')))
    wait.until(EC.visibility_of_element_located((By.XPATH, '//div[@class="lRwqcd"]//span[text()="Kirim"]')))

    radiobuttons = driver.find_elements(By.CLASS_NAME, 'Od2TWd.hYsg7c')
    time.sleep(1)

    for radio in radiobuttons:
        if radio.get_attribute("data-value").lower() == pilihan.lower():
            radio.click()

    bt_kirim = wait.until(EC.element_to_be_clickable((By.XPATH, '//div[@class="lRwqcd"]//span[text()="Kirim"]')))
    bt_kirim.click()

    wait.until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div[2]/div[1]/div/div[3]')))

    another_response = wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div[2]/div[1]/div/div[4]/a')))
    another_response.click()

    time.sleep(1)

workbook = openpyxl.load_workbook("data/electiondata.xlsx")
sheet = workbook['Sheet1']
jumlah_baris = sheet.max_row

for row in sheet.iter_rows(min_row=2, values_only=True):
    nrp, nama, jenis_kelamin, usia, alamat, golongan, posisi, department, divisi, site, pilihan, email = row
    fill_form(email, nrp, nama, jenis_kelamin, usia, alamat, golongan, posisi, department, divisi, site, pilihan)

print("Jumlah data:", jumlah_baris)

print("Selesai")

driver.close()

end_time = time.time()
timer = end_time - start_time

hours = int(timer // 3600)
minutes = int((timer % 3600) // 60)
seconds = int(timer % 60)

print("Waktu eksekusi:", f"{hours} jam, {minutes} menit, {seconds} detik")
