import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service

workbook = openpyxl.load_workbook("data/electiondata.xlsx")
sheet = workbook['Sheet2']

data_to_submit = []

for row in sheet.iter_rows(values_only=True):
    data_to_submit.append(row)

service = Service('chromedriver.exe')
service.start()
driver = webdriver.Chrome(service=service)
driver.get("https://forms.gle/ueXgvWTQRxAb56St7")

# Mendapatkan jumlah baris
rows = sheet.max_row

# Mengisi formulir menggunakan data dari Excel
for r in range(2, rows+1):
    # Membaca data dari kolom pertama (misalnya, kolom 'A')
    nrp = sheet.cell(row=r, column=1).value

    # Identifikasi elemen input dan mengisi dengan data
    driver.find_element(By.CLASS_NAME, "whsOnd.zHQkBf").send_keys(nrp)
    driver.find_element(By.XPATH, '//*[@id="mG61Hd"]/div[2]/div/div[3]/div[1]/div[1]/div').click()

driver.quit()